import os
from random import randint
from pathlib import Path
import torchvision.transforms as T
from torchvision.io import read_image, write_jpeg
import shutil
from PIL import Image
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font

# Chemins des dossiers
folder_path = './images/images_originales'  # Dossier contenant les images originales
output_folder = './images/images_sorties'  # Dossier pour enregistrer les images rognées
output_log_path = './images/text_infos/logs_output.xlsx'

# Vider le dossier contenant les images coupées si nécessaire
if os.path.exists(output_folder):
    shutil.rmtree(output_folder)

# Créer les dossiers nécessaires s'ils n'existent pas
Path(output_folder).mkdir(parents=True, exist_ok=True)

# Liste pour enregistrer les données de sortie pour Excel
log_data = []

# Obtenir tous les fichiers .jpg dans le dossier
image_files = sorted([f for f in os.listdir(folder_path) if f.lower().endswith('.jpg')])

# Si aucune image n'est trouvée, arrête le programme
if len(image_files) == 0:
    print("Aucune image trouvée dans le dossier.")
else:
    # Obtenir le nom de la dernière image (sans extension) pour la première image
    last_image_name, _ = os.path.splitext(image_files[-1])
    new_filename_avant = f"{last_image_name}_cropped.jpg"  # Utilisé pour la première image

    # Parcours de chaque image
    for idx, filename in enumerate(image_files):
        # Extraire le nom de l'image sans l'extension
        original_name, ext = os.path.splitext(filename)

        # Chemin complet de l'image
        img_path = Path(folder_path) / filename

        # Charger l'image
        img = read_image(str(img_path))

        img_PIL = Image.open(img_path)
        image_tensor = T.functional.to_tensor(img_PIL)

        # Obtenir les dimensions de l'image (hauteur et largeur)
        _, height, width = image_tensor.shape

        # Définir une taille de rognage aléatoire
        random_size_w = int(0.1 * width + randint(0, int(0.8 * width)))
        random_size_h = int(0.1 * height + randint(0, int(0.8 * height)))
        crop_size = (random_size_h, random_size_w)  # Taille de rognage
        transform = T.RandomCrop(size=crop_size)

        # Appliquer la transformation de rognage
        cropped_img = transform(img)

        # Créer un nom de fichier pour l'image rognée avec后缀名
        new_filename = f"{original_name}_cropped.jpg"
        new_img_path = Path(output_folder) / new_filename

        # Enregistrer l'image rognée
        write_jpeg(cropped_img, str(new_img_path))

        # Ajouter une ligne au journal pour l'image rognée (sans extension pour l'enregistrement dans Excel)
        log_data.append([original_name, os.path.splitext(new_filename)[0], 1])

        # Ajouter une deuxième ligne pour l'image précédente (sans extension)
        log_data.append([original_name, new_filename_avant, 0])

        # Mettre à jour le nom de la dernière image rognée (sans extension)
        new_filename_avant = os.path.splitext(new_filename)[0]

# Enregistrer les données dans un fichier Excel (sans extension dans le fichier Excel)
log_df = pd.DataFrame(log_data, columns=['Nom original', 'Nom rogné', 'Utilisé l\'image précédente'])
log_df.to_excel(output_log_path, index=False)

# Charger le fichier Excel et appliquer la police Arial à toutes les cellules
wb = load_workbook(output_log_path)
ws = wb.active

# Appliquer la police Arial à chaque cellule
arial_font = Font(name='Arial')
for row in ws.iter_rows():
    for cell in row:
        cell.font = arial_font

# Sauvegarder le fichier avec la nouvelle police
wb.save(output_log_path)

# Copier toutes les images originales dans le dossier de sortie
for file in image_files:
    src_file = os.path.join(folder_path, file)
    dest_file = os.path.join(output_folder, file)
    shutil.copy2(src_file, dest_file)

    print(f"Image {file} copiée dans le dossier de sortie.")
